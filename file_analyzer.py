"""
file_analyzer.py — Astra File & Image Analysis
Handles any file type uploaded via the chat window.

Supported:
  Images    — jpg, png, gif, bmp, webp → describe content
  Text      — txt, md, py, js, sh, yaml, json, csv, log, html, css
  Code      — py, js, ts, java, cpp, c, go, rs, rb, kt, swift, sql
  Documents — pdf (text extraction), docx, xlsx, pptx
  Data      — csv, json, xml, yaml, toml
  Archives  — zip (list contents)
  Any other — reads as text if possible, otherwise describes file info

All analysis is done by llama3 via Ollama locally.
"""

import os
import json
import zipfile
import mimetypes
from pathlib import Path
from llm import call_llm

# ─────────────────────────────────────────────
# FILE TYPE DETECTION
# ─────────────────────────────────────────────

IMAGE_EXTS    = {".jpg",".jpeg",".png",".gif",".bmp",".webp",".tiff",".ico"}
CODE_EXTS     = {".py",".js",".ts",".java",".cpp",".c",".h",".go",".rs",
                 ".rb",".kt",".swift",".php",".cs",".sh",".bash",".ps1",
                 ".sql",".r",".m",".scala",".lua",".pl",".hs",".ex",".clj"}
TEXT_EXTS     = {".txt",".md",".rst",".log",".cfg",".ini",".env",
                 ".html",".htm",".css",".xml",".yaml",".yml",".toml"}
DATA_EXTS     = {".csv",".json",".jsonl",".ndjson",".tsv"}
DOC_EXTS      = {".pdf",".docx",".xlsx",".xls",".pptx",".odt",".ods"}
ARCHIVE_EXTS  = {".zip",".tar",".gz",".7z",".rar"}

MAX_TEXT_CHARS = 8000   # max chars to send to LLM


def get_file_category(path: str) -> str:
    ext = Path(path).suffix.lower()
    if ext in IMAGE_EXTS:   return "image"
    if ext in CODE_EXTS:    return "code"
    if ext in TEXT_EXTS:    return "text"
    if ext in DATA_EXTS:    return "data"
    if ext in DOC_EXTS:     return "document"
    if ext in ARCHIVE_EXTS: return "archive"
    return "unknown"


# ─────────────────────────────────────────────
# EXTRACTORS
# ─────────────────────────────────────────────

def _read_text_file(path: str) -> str:
    """Read a plain text / code file."""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read(MAX_TEXT_CHARS)
        if len(content) == MAX_TEXT_CHARS:
            content += f"\n\n[... truncated at {MAX_TEXT_CHARS} chars ...]"
        return content
    except Exception as e:
        return f"[Could not read file: {e}]"


def _read_csv(path: str) -> str:
    """Read CSV and show summary + first rows."""
    try:
        import csv
        rows = []
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                rows.append(row)
                if i >= 50:
                    rows.append(["... truncated ..."])
                    break
        if not rows:
            return "[Empty CSV file]"
        headers  = rows[0]
        num_rows = len(rows) - 1
        preview  = "\n".join([",".join(r) for r in rows[:6]])
        return (f"CSV file — {len(headers)} columns, ~{num_rows} rows\n"
                f"Columns: {', '.join(headers)}\n\n"
                f"First rows:\n{preview}")
    except Exception as e:
        return _read_text_file(path)


def _read_json(path: str) -> str:
    """Read and pretty-print JSON."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        text = json.dumps(data, indent=2)
        if len(text) > MAX_TEXT_CHARS:
            text = text[:MAX_TEXT_CHARS] + "\n[... truncated ...]"
        return text
    except Exception:
        return _read_text_file(path)


def _read_pdf(path: str) -> str:
    """Extract text from PDF."""
    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages[:20]:    # first 20 pages
                t = page.extract_text()
                if t:
                    text += t + "\n"
                if len(text) > MAX_TEXT_CHARS:
                    text = text[:MAX_TEXT_CHARS] + "\n[... truncated ...]"
                    break
        return text or "[No text found in PDF]"
    except ImportError:
        return "[pdfplumber not installed — run: pip install pdfplumber]"
    except Exception as e:
        return f"[PDF read error: {e}]"


def _read_docx(path: str) -> str:
    """Extract text from Word document."""
    try:
        import docx
        doc  = docx.Document(path)
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        if len(text) > MAX_TEXT_CHARS:
            text = text[:MAX_TEXT_CHARS] + "\n[... truncated ...]"
        return text or "[Empty document]"
    except ImportError:
        return "[python-docx not installed — run: pip install python-docx]"
    except Exception as e:
        return f"[DOCX read error: {e}]"


def _read_xlsx(path: str) -> str:
    """Extract data from Excel file."""
    try:
        import openpyxl
        wb      = openpyxl.load_workbook(path, read_only=True, data_only=True)
        result  = []
        for sheet_name in wb.sheetnames[:5]:   # first 5 sheets
            ws   = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(max_row=50, values_only=True):
                if any(cell is not None for cell in row):
                    rows.append([str(c) if c is not None else "" for c in row])
            if rows:
                result.append(f"Sheet: {sheet_name}")
                result.append("\n".join(["\t".join(r) for r in rows[:10]]))
        return "\n\n".join(result) or "[Empty Excel file]"
    except ImportError:
        return "[openpyxl not installed — run: pip install openpyxl]"
    except Exception as e:
        return f"[XLSX read error: {e}]"


def _read_image(path: str) -> str:
    """Describe image using PIL metadata."""
    try:
        from PIL import Image as PILImage
        img  = PILImage.open(path)
        info = (f"Image file: {Path(path).name}\n"
                f"Format: {img.format}\n"
                f"Size: {img.size[0]}x{img.size[1]} pixels\n"
                f"Mode: {img.mode}")
        # Add EXIF if available
        try:
            exif = img._getexif()
            if exif:
                info += f"\nHas EXIF metadata"
        except Exception:
            pass
        return info
    except Exception as e:
        return f"[Image read error: {e}]"


def _read_archive(path: str) -> str:
    """List contents of zip archive."""
    try:
        with zipfile.ZipFile(path, "r") as z:
            names = z.namelist()
        preview = "\n".join(names[:30])
        if len(names) > 30:
            preview += f"\n... and {len(names)-30} more files"
        return f"ZIP archive — {len(names)} files:\n{preview}"
    except Exception as e:
        return f"[Archive read error: {e}]"


# ─────────────────────────────────────────────
# MAIN ANALYZER
# ─────────────────────────────────────────────

def analyze_file(file_path: str, user_question: str = "") -> str:
    """
    Analyze any file and return Astra's response.
    user_question: optional specific question about the file.
    """
    path     = Path(file_path)
    filename = path.name
    ext      = path.suffix.lower()
    size     = path.stat().st_size if path.exists() else 0
    category = get_file_category(file_path)

    print(f"[FileAnalyzer] Analyzing: {filename} ({category}, {size} bytes)")

    # ── Extract content ──
    if category == "image":
        file_content = _read_image(file_path)
        content_type = "image"
    elif ext == ".csv":
        file_content = _read_csv(file_path)
        content_type = "CSV data"
    elif ext in {".json", ".jsonl"}:
        file_content = _read_json(file_path)
        content_type = "JSON data"
    elif ext == ".pdf":
        file_content = _read_pdf(file_path)
        content_type = "PDF document"
    elif ext == ".docx":
        file_content = _read_docx(file_path)
        content_type = "Word document"
    elif ext in {".xlsx", ".xls"}:
        file_content = _read_xlsx(file_path)
        content_type = "Excel spreadsheet"
    elif ext in ARCHIVE_EXTS:
        file_content = _read_archive(file_path)
        content_type = "archive file"
    elif category in {"code", "text", "data"}:
        file_content = _read_text_file(file_path)
        content_type = "code file" if category == "code" else "text file"
    else:
        # Try reading as text anyway
        file_content = _read_text_file(file_path)
        content_type = "file"

    if not file_content:
        return f"Could not read the file {filename}."

    # ── Build analysis prompt ──
    if user_question:
        task = user_question
    else:
        # Default analysis per type
        if category == "image":
            task = "Describe what this image contains and any metadata."
        elif ext in {".py", ".js", ".ts", ".java"}:
            task = ("Analyse this code. Explain what it does, "
                    "identify any bugs or issues, and suggest improvements.")
        elif ext == ".csv":
            task = ("Analyse this data. Describe the structure, "
                    "key columns, data quality, and any interesting patterns.")
        elif ext in {".json", ".yaml", ".yml"}:
            task = ("Analyse this configuration/data file. "
                    "Explain its structure and purpose.")
        elif ext in {".pdf", ".docx"}:
            task = ("Summarise this document. "
                    "Extract key points and important information.")
        elif ext in {".xlsx", ".xls"}:
            task = ("Analyse this spreadsheet. "
                    "Describe the data, structure, and key insights.")
        elif ext in {".sh", ".bash", ".ps1"}:
            task = ("Analyse this script. "
                    "Explain what it does and any potential issues.")
        else:
            task = ("Analyse this file. "
                    "Explain its content, purpose, and key information.")

    prompt = f"""You are Astra, an AI assistant helping Sowmik analyse a file.

File name: {filename}
File type: {content_type}
File size: {size:,} bytes

File content:
{file_content}

Task: {task}

Provide a clear, helpful analysis. Be specific and practical.
If it is code, identify the language and explain the logic.
If it is data, give statistics and insights.
If it is a document, summarise the key points.
Address Sowmik directly in your response."""

    response = call_llm(prompt, max_tokens=1000)
    return response


def get_file_icon(path: str) -> str:
    """Return an emoji icon for the file type."""
    ext      = Path(path).suffix.lower()
    category = get_file_category(path)
    icons = {
        "image":    "🖼️",
        "code":     "💻",
        "text":     "📄",
        "data":     "📊",
        "document": "📋",
        "archive":  "📦",
    }
    ext_icons = {
        ".pdf":  "📕",
        ".docx": "📘",
        ".xlsx": "📗",
        ".pptx": "📙",
        ".csv":  "📊",
        ".json": "🔧",
        ".yaml": "🔧",
        ".py":   "🐍",
        ".js":   "🌐",
        ".sh":   "⚙️",
        ".zip":  "📦",
        ".png":  "🖼️",
        ".jpg":  "🖼️",
    }
    return ext_icons.get(ext, icons.get(category, "📎"))
